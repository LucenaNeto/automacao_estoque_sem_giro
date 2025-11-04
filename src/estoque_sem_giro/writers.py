from __future__ import annotations

from pathlib import Path
import csv
import re, os
from .config import Config, yesterday_str
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def write_consolidated_csv(records: list[dict], cfg: Config) -> Path:
    if not records:
        raise ValueError("Nenhum registro válido para salvar.")
    records = sorted(records, key=lambda r: (r.get("PDV",""), r.get("SKU","")))
    cfg.output_dir.mkdir(parents=True, exist_ok=True)
    out = cfg.output_dir / f"{cfg.output_basename}_{yesterday_str(cfg)}.csv"
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cfg.final_fields)
        w.writeheader()
        for rec in records:
            w.writerow({k: rec.get(k, "") for k in cfg.final_fields})
    return out

def write_csvs_by_pdv(records: list[dict], cfg: Config) -> dict[str, Path]:
    if not records:
        return {}
    date_str = yesterday_str(cfg)
    folder = cfg.output_dir / f"por_pdv_{date_str}"
    folder.mkdir(parents=True, exist_ok=True)

    groups: dict[str, list[dict]] = {}
    for rec in records:
        pdv = (rec.get("PDV") or "").strip() or "SEM_PDV"
        pdv = re.sub(r"[^\w\-]+", "_", pdv)
        groups.setdefault(pdv, []).append(rec)

    paths: dict[str, Path] = {}
    for pdv, rows in sorted(groups.items(), key=lambda kv: kv[0]):
        rows_sorted = sorted(rows, key=lambda r: (r.get("SKU","")))
        path = folder / f"{cfg.output_basename}_{date_str}_PDV_{pdv}.csv"
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=cfg.final_fields)
            w.writeheader()
            for rec in rows_sorted:
                w.writerow({k: rec.get(k, "") for k in cfg.final_fields})
        paths[pdv] = path
    return paths


def write_reports_xlsx_by_pdv(records: list[dict], discontinued: list[dict], cfg: Config) -> dict[str, Path]:
    """
    Gera um .xlsx por PDV em:
      data/output/relatorio_finalizado_DD_MM_AAAA/relatorio_finalizado_DD_MM_AAAA_PDV_<pdv>.xlsx

    Abas:
      - cfg.report_sheet_main (Estoque sem Giro)  -> usa cfg.final_fields
      - cfg.report_sheet_disc (Descontinuados)    -> usa cfg.discontinued_fields

    Layout:
      - Linha 1: título "Grupo Ana Sobral" (mesclado A1:ÚltimaColuna1)
      - Linha 2: LOGO (se cfg.logo_path existir)
      - Linha 3: vazia
      - Linha 4: cabeçalho
      - Dados a partir da linha 5
      - Coluna "CURVA" na aba principal: A/B=verde, C=amarelo, D/E=vermelho
    """
    if records is None:
        records = []
    if discontinued is None:
        discontinued = []

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    import os, re, math

    GROUP_NAME = "Grupo Ana Sobral"
    date_str = yesterday_str(cfg)
    folder = cfg.output_dir / f"{cfg.report_folder_prefix}_{date_str}"
    folder.mkdir(parents=True, exist_ok=True)

    # --- helper: inserir logo (A2), escalando para largura/altura máximas ---
    def _insert_logo(ws, ncols: int):
        if not cfg.logo_path or not cfg.logo_path.exists():
            # sem logo, só garante a altura da linha 2 p/ futuro
            ws.row_dimensions[2].height = cfg.logo_row_height
            return
        try:
            img = XLImage(str(cfg.logo_path))
        except Exception:
            ws.row_dimensions[2].height = cfg.logo_row_height
            return

        # calcula escala para caber: largura máx e altura máx (altura em pontos -> px aprox.)
        max_w = max(40, int(cfg.logo_max_width_px))
        max_h_px = int(max(24, cfg.logo_row_height) * 96 / 72)  # 1pt ≈ 96/72 px
        scale = min(max_w / max(1, img.width), max_h_px / max(1, img.height))
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)

        ws.row_dimensions[2].height = cfg.logo_row_height
        ws.add_image(img, "A2")  # ancora no canto esquerdo; (centralização real exige desenho mais complexo)

    # --- Agrupar por PDV (sanitizado para nome de arquivo) ---
    def sanitize_pdv(v: str) -> str:
        v = (v or "").strip() or "SEM_PDV"
        return re.sub(r"[^\w\-]+", "_", v)

    groups_main: dict[str, list[dict]] = {}
    for rec in records:
        groups_main.setdefault(sanitize_pdv(rec.get("PDV", "")), []).append(rec)

    groups_disc: dict[str, list[dict]] = {}
    for rec in discontinued:
        groups_disc.setdefault(sanitize_pdv(rec.get("PDV", "")), []).append(rec)

    all_pdvs = sorted(set(groups_main.keys()) | set(groups_disc.keys()))

    # Larguras sugeridas por coluna
    widths_main = {
        "PDV": 12, "SKU": 14, "DESCRIÇÃO": 50, "MARCA": 16, "CURVA": 10, "ESTOQUE_ATUAL": 18
    }
    disc_header = list(cfg.discontinued_fields)
    widths_disc = {
        "PDV": 12, "SKU": 14, "SKU_PARA": 18, "DESCRIÇÃO": 50,
        "ESTOQUE ATUAL": 18, "FASES DO PRODUTO": 24, "MARCA": 14,
    }

    # Cores CURVA
    FILL_GREEN  = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
    FILL_YELLOW = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
    FILL_RED    = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")

    out_paths: dict[str, Path] = {}

    for pdv in all_pdvs:
        rows_main = sorted(groups_main.get(pdv, []), key=lambda r: (r.get("SKU", "")))
        rows_disc = sorted(groups_disc.get(pdv, []), key=lambda r: (r.get("SKU", "")))

        # === Workbook por PDV ===
        wb = Workbook()

        # ---------- ABA PRINCIPAL ----------
        ws = wb.active
        ws.title = cfg.report_sheet_main
        main_header = list(cfg.final_fields)
        ncols_main = len(main_header)
        last_main_col = get_column_letter(ncols_main)

        # topo visual
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_main)
        cell_title = ws.cell(row=1, column=1, value=GROUP_NAME)
        cell_title.font = Font(bold=True, size=16)
        cell_title.alignment = Alignment(horizontal="center", vertical="center")

        # LOGO na linha 2
        _insert_logo(ws, ncols_main)

        # cabeçalho (linha 4)
        header_row = 4
        for i, col_name in enumerate(main_header, start=1):
            c = ws.cell(row=header_row, column=i, value=col_name)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(i)].width = widths_main.get(col_name, 16)

        ws.freeze_panes = f"A{header_row+1}"
        ws.auto_filter.ref = f"A{header_row}:{last_main_col}{header_row}"

        # dados (linha 5+)
        first_data_row = header_row + 1
        for r_idx, rec in enumerate(rows_main, start=first_data_row):
            for c_idx, col_name in enumerate(main_header, start=1):
                ws.cell(row=r_idx, column=c_idx, value=rec.get(col_name, ""))

        # coloração CURVA
        if "CURVA" in main_header:
            curva_col_idx = main_header.index("CURVA") + 1
            for r in range(first_data_row, ws.max_row + 1):
                v = str(ws.cell(row=r, column=curva_col_idx).value or "").strip().upper()
                if v in {"A", "B"}:
                    ws.cell(row=r, column=curva_col_idx).fill = FILL_GREEN
                elif v == "C":
                    ws.cell(row=r, column=curva_col_idx).fill = FILL_YELLOW
                elif v in {"D", "E"}:
                    ws.cell(row=r, column=curva_col_idx).fill = FILL_RED

        # ---------- ABA DESCONTINUADOS ----------
        ws2 = wb.create_sheet(cfg.report_sheet_disc)
        ncols_disc = len(disc_header)
        last_disc_col = get_column_letter(ncols_disc)

        # topo visual
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_disc)
        cell_title2 = ws2.cell(row=1, column=1, value=GROUP_NAME)
        cell_title2.font = Font(bold=True, size=16)
        cell_title2.alignment = Alignment(horizontal="center", vertical="center")

        # LOGO na linha 2 da aba de descontinuados
        _insert_logo(ws2, ncols_disc)

        # cabeçalho (linha 4)
        for i, col_name in enumerate(disc_header, start=1):
            c = ws2.cell(row=header_row, column=i, value=col_name)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws2.column_dimensions[get_column_letter(i)].width = widths_disc.get(col_name, 16)

        ws2.freeze_panes = f"A{header_row+1}"
        ws2.auto_filter.ref = f"A{header_row}:{last_disc_col}{header_row}"

        # dados (linha 5+)
        first_disc_row = header_row + 1
        for r_idx, rec in enumerate(rows_disc, start=first_disc_row):
            for c_idx, col_name in enumerate(disc_header, start=1):
                ws2.cell(row=r_idx, column=c_idx, value=rec.get(col_name, ""))

        # salvar atômico
        path = folder / f"{cfg.report_folder_prefix}_{date_str}_PDV_{pdv}.xlsx"
        tmp  = path.with_suffix(path.suffix + ".tmp")
        wb.save(tmp)
        os.replace(tmp, path)
        out_paths[pdv] = path

    return out_paths


#DESCONTINUADOS 

def write_discontinued_csvs_by_pdv(records: list[dict], cfg: Config) -> dict[str, Path]:
    """
    Gera CSVs de DESCONTINUADOS por PDV em:
      data/output/descontinuados_DD_MM_AAAA/descontinuados_DD_MM_AAAA_PDV_<pdv>.csv

    Campos (ordem exata): cfg.discontinued_fields
    """
    if not records:
        return {}

    
    date_str = yesterday_str(cfg)
    folder = cfg.output_dir / f"{cfg.discontinued_folder_prefix}_{date_str}"
    folder.mkdir(parents=True, exist_ok=True)

    # agrupar por PDV
    groups: dict[str, list[dict]] = {}
    for rec in records:
        pdv = (rec.get("PDV") or "").strip() or "SEM_PDV"
        pdv = re.sub(r"[^\w\-]+", "_", pdv)
        groups.setdefault(pdv, []).append(rec)

    paths: dict[str, Path] = {}
    for pdv, rows in sorted(groups.items(), key=lambda kv: kv[0]):
        rows_sorted = sorted(rows, key=lambda r: (r.get("SKU", "")))
        path = folder / f"{cfg.discontinued_folder_prefix}_{date_str}_PDV_{pdv}.csv"
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=cfg.discontinued_fields)
            w.writeheader()
            for rec in rows_sorted:
                w.writerow({k: rec.get(k, "") for k in cfg.discontinued_fields})
        paths[pdv] = path

    return paths
