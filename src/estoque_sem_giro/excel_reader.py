from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from typing import Iterable

_HEADER_TOKENS = {
    "sku", "descrição", "descricao", "curva", "classe",
    "pdv", "estoque", "estoque atual", "estoque_atual"
}

def _clean_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "none"):
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    return s

def looks_like_header(values: list[str]) -> bool:
    hits = sum(1 for v in values if v.lower().strip() in _HEADER_TOKENS)
    return hits >= 2

def open_workbook(path: Path, data_only: bool = True):
    return load_workbook(filename=str(path), read_only=True, data_only=data_only)

def find_header_row(ws, cols: Iterable[str], search_limit: int = 200) -> int | None:
    idxs = [column_index_from_string(c) for c in cols]   # A=1, C=3, E=5, I=9, J=10
    max_col = max(idxs)
    max_row_scan = min(ws.max_row or 1, search_limit)

    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row_scan, min_col=1, max_col=max_col, values_only=True),
        start=1
    ):
        values = []
        for i in idxs:
            v = row[i-1] if len(row) >= i else None
            values.append(_clean_str(v))
        if looks_like_header(values):
            return r_idx
    return None

def preview_sheet(ws, cols: Iterable[str], max_rows: int) -> list[list[str]]:
    """Retorna até max_rows de amostras como lista de linhas de strings."""
    idxs = [column_index_from_string(c) for c in cols]
    max_col = max(idxs)
    samples: list[list[str]] = []
    for row in ws.iter_rows(min_row=1, min_col=1, max_col=max_col, values_only=True):
        vals = []
        has_any = False
        for i in idxs:
            v = row[i-1] if len(row) >= i else None
            s = _clean_str(v)
            if s:
                has_any = True
            vals.append(s)
        if has_any:
            samples.append(vals)
            if len(samples) >= max_rows:
                break
    return samples
