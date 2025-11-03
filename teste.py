from __future__ import annotations

import re
from pathlib import Path
import logging
from typing import Optional, Iterable
from zoneinfo import ZoneInfo
from datetime import datetime, timedelta
import csv
import shutil

# ---------- Config ----------
PROJECT_ROOT = Path(__file__).parent.resolve()
DATA_DIR     = PROJECT_ROOT / "data"
INPUT_DIR    = DATA_DIR / "input"
OUTPUT_DIR   = DATA_DIR / "output"
ARCHIVE_DIR  = DATA_DIR / "archived"

EXPECTED_SHEETS: tuple[str, ...] = ("EUD", "BOT", "QDB")
PREVIEW_COLS: tuple[str, ...]    = ("A", "C", "E", "I", "J")  # A=SKU, C=DESCRIÇÃO, E=CURVA/CLASSE, I=PDV, J=ESTOQUE_ATUAL
TZ = ZoneInfo("America/Recife")  # use "America/Maceio" se preferir

FINAL_FIELDS = ("PDV", "SKU", "DESCRIÇÃO", "MARCA", "CURVA", "ESTOQUE_ATUAL")
OUTPUT_BASENAME = "Estoque_sem_giro"

# ---------- Infra básica ----------
def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
    )

def ensure_dirs() -> None:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)

def is_excel(p: Path) -> bool:
    # aceita .xlsx (ignora temp do Excel começando com ~$
    return p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")

def list_excels() -> list[Path]:
    if not INPUT_DIR.exists():
        return []
    files = [p for p in INPUT_DIR.iterdir() if is_excel(p)]
    files.sort(key=lambda x: (x.stat().st_mtime, x.name))
    return files

def latest_excel() -> Optional[Path]:
    files = list_excels()
    return files[-1] if files else None

# ---------- Datas ----------
def yesterday_str_recife() -> str:
    dt = datetime.now(TZ) - timedelta(days=1)
    return dt.strftime("%d_%m_%Y")

# ---------- Util (limpeza e detecção de cabeçalho) ----------
def _clean_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "none"):
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    return s

_HEADER_TOKENS = {
    "sku", "descrição", "descricao", "curva", "classe",
    "pdv", "estoque", "estoque atual", "estoque_atual"
}

def _looks_like_header(values: list[str]) -> bool:
    hits = 0
    for v in values:
        vv = v.lower().strip()
        if vv in _HEADER_TOKENS:
            hits += 1
    return hits >= 2  # pelo menos 2 termos de cabeçalho

def find_header_row(ws, cols=("A","C","E","I","J"), search_limit=200) -> int | None:
    """
    Localiza a linha de cabeçalho analisando SOMENTE as colunas-alvo,
    usando leitura streaming (iter_rows) — muito mais rápido em read_only.
    """
    from openpyxl.utils.cell import column_index_from_string

    # letras -> índices (A=1, C=3, E=5, I=9, J=10)
    idxs = [column_index_from_string(c) for c in cols]
    max_col = max(idxs)
    max_row_scan = min(ws.max_row or 1, search_limit)

    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row_scan, min_col=1, max_col=max_col, values_only=True),
        start=1
    ):
        values = []
        for i in idxs:
            v = row[i-1] if len(row) >= i else None
            values.append(_clean_str(v))
        if _looks_like_header(values):
            return r_idx
    return None


# ---------- Preview visual (opcional) ----------
def _format_row(values: list[str]) -> str:
    widths = [max(12, min(40, max(len(v), 4))) for v in values]
    return " | ".join(v.ljust(w)[:w] for v, w in zip(values, widths))

def preview_sheet(ws, sheet_name: str, cols: Iterable[str] = PREVIEW_COLS, max_rows: int = 5) -> None:
    from openpyxl.utils.cell import column_index_from_string
    logging.info("Prévia da aba: %s (colunas: %s)", sheet_name, ", ".join(cols))
    header = [f"{c}" for c in cols]
    logging.info(_format_row(header))
    printed = 0
    for row_idx in range(1, ws.max_row + 1):
        row_vals: list[str] = []
        has_any = False
        for c in cols:
            col_idx = column_index_from_string(c)
            cell = ws.cell(row=row_idx, column=col_idx)
            val = _clean_str(cell.value)
            if val:
                has_any = True
            row_vals.append(val)
        if has_any:
            logging.info(_format_row(row_vals))
            printed += 1
        if printed >= max_rows:
            break
    if printed == 0:
        logging.info("(sem conteúdo visível nessas colunas)")

# ---------- Abertura workbook ----------
def open_workbook(xlsx_path: Path):
    from openpyxl import load_workbook
    return load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)

# ---------- Extração normalizada ----------
_COL_MAP = {
    "A": "SKU",
    "C": "DESCRIÇÃO",
    "E": "CURVA",         # no seu arquivo pode vir como "CLASSE"; guardamos em CURVA
    "I": "PDV",
    "J": "ESTOQUE_ATUAL",
}

def extract_records_from_sheet(ws, marca: str) -> list[dict]:
    """
    Extrai colunas A,C,E,I,J como:
    { PDV, SKU, DESCRIÇÃO, MARCA, CURVA, ESTOQUE_ATUAL }
    Lê com iter_rows(values_only=True) para performance,
    começa logo abaixo do cabeçalho e pula linhas sem PDV/SKU.
    """
    # mapeamento fixo de posições (A=1, C=3, E=5, I=9, J=10)
    col_pos = {"A": 1, "C": 3, "E": 5, "I": 9, "J": 10}

    # achar cabeçalho com a função otimizada
    header_row = find_header_row(ws, cols=("A","C","E","I","J"))
    start_row = (header_row + 1) if header_row else 1
    max_col = max(col_pos.values())

    records: list[dict] = []
    for row in ws.iter_rows(min_row=start_row, min_col=1, max_col=max_col, values_only=True):
        # pegar valores com segurança — se a linha tiver menos colunas, tratamos como vazio
        def at(i: int) -> str:
            return _clean_str(row[i-1] if len(row) >= i else "")

        sku = at(col_pos["A"])
        desc = at(col_pos["C"])
        curva = at(col_pos["E"])   # no seu arquivo pode vir “CLASSE”; guardamos em CURVA
        pdv  = at(col_pos["I"])
        est  = at(col_pos["J"])

        # pular totalmente vazias
        if not any((sku, desc, curva, pdv, est)):
            continue
        # requisitos mínimos
        if not sku or not pdv:
            continue

        records.append({
            "PDV":           pdv,
            "SKU":           sku,
            "DESCRIÇÃO":     desc,
            "MARCA":         marca,
            "CURVA":         curva,
            "ESTOQUE_ATUAL": est,
        })

    return records


def extract_all(wb) -> list[dict]:
    data: list[dict] = []
    for sheet in EXPECTED_SHEETS:
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            recs = extract_records_from_sheet(ws, marca=sheet)
            logging.info("[Resumo] %s -> %d registros", sheet, len(recs))
            data.extend(recs)
        else:
            logging.warning("Aba esperada não encontrada: %s", sheet)
    return data

# ---------- Escrita CSV ----------
def write_consolidated_csv(records: list[dict], output_dir: Path, base_name: str = OUTPUT_BASENAME) -> Path:
    if not records:
        raise ValueError("Nenhum registro válido para salvar.")

    # Agrupa por PDV e organiza por SKU dentro do PDV
    try:
        records = sorted(records, key=lambda r: (r.get("PDV", ""), r.get("SKU", "")))
    except Exception:
        pass

    output_dir.mkdir(parents=True, exist_ok=True)
    fname = f"{base_name}_{yesterday_str_recife()}.csv"
    out_path = output_dir / fname

    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FINAL_FIELDS)
        writer.writeheader()
        for rec in records:
            writer.writerow({
                "PDV":           rec.get("PDV", ""),
                "SKU":           rec.get("SKU", ""),
                "DESCRIÇÃO":     rec.get("DESCRIÇÃO", ""),
                "MARCA":         rec.get("MARCA", ""),
                "CURVA":         rec.get("CURVA", ""),
                "ESTOQUE_ATUAL": rec.get("ESTOQUE_ATUAL", ""),
            })

    logging.info("[OK] CSV consolidado salvo (agrupado por PDV): %s (linhas=%d)", out_path, len(records))
    return out_path

def write_csvs_by_pdv(records: list[dict], output_dir: Path, base_name: str = OUTPUT_BASENAME) -> dict[str, Path]:
    """
    Gera 1 CSV por PDV dentro de data/output/por_pdv_DD_MM_AAAA.
    Retorna {pdv: Path}.
    """
    if not records:
        return {}

    # Agrupa por PDV e ordena por SKU dentro do PDV
    groups: dict[str, list[dict]] = {}
    for rec in records:
        pdv = str(rec.get("PDV", "")).strip() or "SEM_PDV"
        # sanitiza para nome de arquivo
        pdv = re.sub(r"[^\w\-]+", "_", pdv)
        groups.setdefault(pdv, []).append(rec)

    date_str = yesterday_str_recife()
    folder = output_dir / f"por_pdv_{date_str}"
    folder.mkdir(parents=True, exist_ok=True)

    paths: dict[str, Path] = {}
    for pdv, rows in sorted(groups.items(), key=lambda kv: kv[0]):
        rows_sorted = sorted(rows, key=lambda r: (r.get("SKU", "")))
        path = folder / f"{base_name}_{date_str}_PDV_{pdv}.csv"
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=FINAL_FIELDS)
            writer.writeheader()
            for r in rows_sorted:
                writer.writerow({
                    "PDV":           r.get("PDV", ""),
                    "SKU":           r.get("SKU", ""),
                    "DESCRIÇÃO":     r.get("DESCRIÇÃO", ""),
                    "MARCA":         r.get("MARCA", ""),
                    "CURVA":         r.get("CURVA", ""),
                    "ESTOQUE_ATUAL": r.get("ESTOQUE_ATUAL", ""),
                })
        paths[pdv] = path

    logging.info("[OK] %d arquivos por PDV salvos em: %s", len(paths), folder)
    return paths


# ---------- Arquivar XLSX ----------
def archive_xlsx(xlsx: Path) -> Path:
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = ARCHIVE_DIR / f"{xlsx.stem}__processado_{ts}{xlsx.suffix}"
    shutil.move(str(xlsx), str(dest))
    logging.info("[OK] Arquivo original arquivado em: %s", dest)
    return dest

# ---------- Debug auxiliar ----------
def debug_sheet(ws, max_cols=40, max_rows=5):
    from openpyxl.utils import get_column_letter
    print(f"\n=== Debug: {ws.title} ===")
    print("dimension:", ws.calculate_dimension(), "| max_row:", ws.max_row, "| max_col:", ws.max_column)
    for r in range(1, min(ws.max_row, max_rows) + 1):
        row_vals = []
        has_any = False
        for c in range(1, min(ws.max_column, max_cols) + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append("" if v is None else str(v))
            if v not in (None, ""):
                has_any = True
        if has_any:
            cols = [get_column_letter(i) for i in range(1, len(row_vals)+1)]
            print(" ".join(cols))
            print(" | ".join(x if x != "" else "." for x in row_vals))
            break

# ---------- Fluxo principal ----------
def main() -> int:
    setup_logging()
    ensure_dirs()

    files = list_excels()
    if not files:
        logging.error("Nenhum .xlsx encontrado. Coloque seu arquivo em: %s", INPUT_DIR)
        return 2

    xlsx = latest_excel()
    logging.info("Arquivo selecionado (mais recente): %s", xlsx.name)

    # 1ª tentativa: data_only=True (pega valores calculados)
    try:
        wb = open_workbook(xlsx)
    except Exception as e:
        logging.exception("Falha ao abrir o arquivo com openpyxl: %s", e)
        return 3

    logging.info("Abas disponíveis: %s", ", ".join(wb.sheetnames))

    # Prévia curta (pode comentar se preferir)
    for sheet in EXPECTED_SHEETS:
        if sheet in wb.sheetnames:
            preview_sheet(wb[sheet], sheet_name=sheet, cols=PREVIEW_COLS, max_rows=3)

    # Extrair
    records = extract_all(wb)
    try:
        wb.close()
    except Exception:
        pass

    # Fallback diagnóstico: formulas sem cache
    if not records:
        logging.warning("Nenhum registro extraído. Tentando fallback com data_only=False (diagnóstico de fórmulas).")
        try:
            from openpyxl import load_workbook
            wb2 = load_workbook(filename=str(xlsx), read_only=True, data_only=False)
            records = extract_all(wb2)
            try:
                wb2.close()
            except Exception:
                pass
        except Exception as e:
            logging.exception("Falha no fallback data_only=False: %s", e)

    if not records:
        logging.error(
            "Ainda sem registros. Possível planilha com FÓRMULAS/consultas sem cache.\n"
            "- Abra no Excel, pressione Ctrl+Alt+F9 (recalcular tudo) e salve.\n"
            "- Depois rode o script novamente."
        )
        return 4

    out_path = write_consolidated_csv(records, OUTPUT_DIR, base_name=OUTPUT_BASENAME)

# Gera CSVs por PDV 
    write_csvs_by_pdv(records, OUTPUT_DIR, base_name=OUTPUT_BASENAME)

    # Arquiva o arquivo original
    try:
        archive_xlsx(xlsx)
    except Exception as e:
        logging.warning("Não foi possível arquivar %s: %s", xlsx.name, e)

    logging.info("Concluído. Arquivo gerado: %s", out_path)
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
